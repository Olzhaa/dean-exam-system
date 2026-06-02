"""SDU BS exam schedule Excel I/O.

Input/output format (Лист1, 13 columns):
  Row 0-7: header/title rows (skipped on import, reproduced on export)
  Row 8:   column headers
  Row 9+:  data rows

Columns:
  0  Курс              -> course_year
  1  Мерзімі           -> exam_date
  2  Уақыты            -> exam_time (start) + computed duration from range
  3  Емтихан қабылдаушы -> lecturer
  4  Бақылаушылар / Proctor -> proctors text (FILLED IN ON EXPORT)
  5  Пән коды          -> course_code
  6  Пән атауы         -> course_name
  7  БББ атауы         -> program_name
  8  ECTS              -> ects
  9  Студент саны      -> student_count
  10 Аудитория         -> room_number (raw text, may include multiple rooms)
  11 Емтихан форматы   -> exam_format
  12 Емтихан ұзақтығы  -> duration (mins, parsed from "60 мин")
"""
from __future__ import annotations
from datetime import datetime, time, date
from typing import Optional, List, Tuple
import io
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from sqlalchemy.orm import Session

from . import models


# Row 8 is 0-indexed; in Excel it's row 9
HEADER_ROW_IDX = 8
DATA_START_IDX = 9
EXPECTED_COLS = 13


def _parse_time_range(s: str) -> Tuple[Optional[time], Optional[int]]:
    """Parse '09.00 - 10.00' or '09:00 - 10:30' -> (start_time, duration_min)."""
    if not s or not isinstance(s, str):
        return None, None
    cleaned = s.replace(".", ":").strip()
    m = re.match(r"(\d{1,2}):(\d{2})\s*[-–—]\s*(\d{1,2}):(\d{2})", cleaned)
    if not m:
        return None, None
    h1, m1, h2, m2 = map(int, m.groups())
    start = time(h1, m1)
    duration = (h2 * 60 + m2) - (h1 * 60 + m1)
    return start, max(duration, 0)


def _parse_duration(s) -> Optional[int]:
    if s is None:
        return None
    s = str(s).strip()
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def _parse_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None


def _norm_rooms(raw) -> str:
    """Normalize a multi-room cell into a comma-separated, deduped string."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return ""
    parts = re.split(r"[,;]\s*|\s{2,}|\n", s)
    rooms = []
    seen = set()
    for p in parts:
        p = p.strip()
        if p and p not in seen:
            rooms.append(p)
            seen.add(p)
    return ", ".join(rooms)


def _safe_str(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip().replace("\n", " ")


def _safe_int(v) -> Optional[int]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def import_bs_excel(db: Session, content: bytes) -> Tuple[int, int, List[str]]:
    """Parse SDU BS exam schedule and persist Exam rows. Returns (added, skipped, errors)."""
    df = pd.read_excel(io.BytesIO(content), header=None)
    if df.shape[1] < EXPECTED_COLS:
        raise ValueError(f"Күтілген ${EXPECTED_COLS} баған, табылды: {df.shape[1]}")

    added = 0
    skipped = 0
    errors: List[str] = []

    for i in range(DATA_START_IDX, len(df)):
        row = df.iloc[i]
        course_code = _safe_str(row[5])
        if not course_code:
            skipped += 1
            continue
        try:
            exam_date_v = _parse_date(row[1])
            start_time, dur_from_range = _parse_time_range(_safe_str(row[2]))
            if not exam_date_v or not start_time:
                skipped += 1
                continue
            dur_raw = _parse_duration(row[12])
            duration = dur_raw or dur_from_range or 90

            rooms = _norm_rooms(row[10])
            student_count = _safe_int(row[9]) or 0
            ects = _safe_int(row[8])

            # required_proctors per room: default 2; if no rooms (online project), 0
            n_rooms = len([r for r in rooms.split(",") if r.strip()]) if rooms else 0
            req_proctors_per_room = 2 if n_rooms > 0 else 0

            # check duplicate (same course + date + time + lecturer)
            existing = db.query(models.Exam).filter(
                models.Exam.course_code == course_code,
                models.Exam.exam_date == exam_date_v,
                models.Exam.exam_time == start_time,
                models.Exam.lecturer == _safe_str(row[3]),
            ).first()
            if existing:
                skipped += 1
                continue

            exam = models.Exam(
                course_code=course_code,
                course_name=_safe_str(row[6]),
                program_name=_safe_str(row[7]),
                lecturer=_safe_str(row[3]),
                course_year=_safe_str(row[0]),
                ects=ects,
                student_count=student_count,
                exam_format=_safe_str(row[11]),
                duration=duration,
                room_number=rooms,
                required_proctors=req_proctors_per_room,
                exam_date=exam_date_v,
                exam_time=start_time,
            )
            db.add(exam)
            added += 1
        except Exception as e:
            errors.append(f"Жол {i+1}: {e}")
    db.commit()
    return added, skipped, errors[:20]


def export_bs_excel(db: Session) -> bytes:
    """Generate BS-format Excel with proctors filled in column 4."""
    exams = db.query(models.Exam).order_by(models.Exam.exam_date, models.Exam.exam_time).all()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # Title block (rows 1-8 in Excel = idx 0-7)
    ws.cell(row=2, column=12, value='"БЕКІТЕМІН"\nОқу істері жөніндегі\nпроректор\n_______________Богданчиков А.\n"____"____________2025ж.').alignment = Alignment(wrap_text=True)
    ws.cell(row=3, column=1, value="SDU БИЗНЕС МЕКТЕБІ")
    ws.cell(row=4, column=1, value="6В04101 - ЭКОНОМИКА, 6В04102 - МЕНЕДЖМЕНТ, 6В04103 - ЕСЕП ЖӘНЕ АУДИТ,")
    ws.cell(row=5, column=1, value="6В04104 - ҚАРЖЫ, 6В04105 - ДИДЖИТАЛ МАРКЕТИНГ БІЛІМ БЕРУ БАҒДАРЛАМАЛАРЫ")
    ws.cell(row=6, column=1, value="АРАЛЫҚ АТТЕСТАТТАУ КЕСТЕСІ")
    ws.cell(row=7, column=1, value="2024-2025 оқу жылы, көктемгі семестр")

    headers = [
        "Курс", "Мерзімі", "Уақыты", "Емтихан қабылдаушы",
        "Бақылаушылар / Proctor", "Пән коды", "Пән атауы",
        "БББ атауы", "ECTS", "Студент саны",
        "Аудитория / Платфо*рма", "Емтихан форматы тест/проект, жазбаша/ ауызша",
        "Емтихан ұзақтығы",
    ]
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill = PatternFill("solid", fgColor="E5E7EB")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=col_idx, value=h)
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border
        c.fill = fill

    # Data rows
    for idx, ex in enumerate(exams, start=10):
        # build proctors text per room
        proctor_lines = []
        # group assignments by room
        by_room = {}
        for a in ex.assignments:
            r = a.room or ""
            by_room.setdefault(r, []).append(a.employee.name)
        # preserve room order from rooms_list
        for r in ex.rooms_list:
            for name in by_room.get(r, []):
                proctor_lines.append(f"{name} ({r})")
        # any leftover (no room)
        for r, names in by_room.items():
            if r not in ex.rooms_list:
                for name in names:
                    proctor_lines.append(name)
        proctors_text = "\n".join(proctor_lines) if proctor_lines else (
            "Қажет емес" if ex.required_proctors == 0 else ""
        )

        time_str = ""
        if ex.exam_time:
            start = ex.exam_time
            end_min = (start.hour * 60 + start.minute) + (ex.duration or 0)
            eh, em = divmod(end_min, 60)
            time_str = f"{start.hour:02d}.{start.minute:02d} - {eh:02d}.{em:02d}"

        row_vals = [
            ex.course_year or "",
            ex.exam_date.strftime("%Y-%m-%d") if ex.exam_date else "",
            time_str,
            ex.lecturer or "",
            proctors_text,
            ex.course_code or "",
            ex.course_name or "",
            ex.program_name or "",
            ex.ects if ex.ects is not None else "",
            ex.student_count or "",
            ex.room_number or "",
            ex.exam_format or "",
            f"{ex.duration} мин" if ex.duration else "",
        ]
        for col_idx, v in enumerate(row_vals, start=1):
            c = ws.cell(row=idx, column=col_idx, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border

    # Column widths
    widths = [10, 12, 14, 24, 32, 14, 24, 28, 6, 10, 22, 18, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=9, column=col_idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
