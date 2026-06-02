"""FX exam I/O: import raw registration list, export 2-sheet BS-style schedule.

INPUT: FX.xlsx with columns:
  №, FACULTY, DEPARTMENT, EDU_LEVEL, CIPHER, SPECIALITY, EDU_LANG, PERIOD_COUNT,
  COURSE_CODE, COURSE_TITLE, CREDITS, ECTS, SECTION, INSTRUCTOR, YEAR, TERM,
  EXAM_TYPE, STUD_ID, STUD_FULL_NAME

OUTPUT: 2-sheet xlsx
  Sheet "FX кестесі": exam slots (one per session)
  Sheet "Өтініш берген білімгерлер тізім": one row per student-course assignment
"""
from __future__ import annotations
from typing import List, Tuple
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from sqlalchemy.orm import Session

from . import models


def _safe_str(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def _safe_int(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def import_fx_excel(db: Session, content: bytes) -> Tuple[int, int, List[str]]:
    """Import raw FX registration sheet."""
    df = pd.read_excel(io.BytesIO(content))
    df.columns = [str(c).strip().upper() for c in df.columns]

    required = {"COURSE_CODE", "STUD_ID", "STUD_FULL_NAME"}
    if not required.issubset(df.columns):
        raise ValueError(f"Қажетті бағандар: {', '.join(required)}")

    added = 0
    skipped = 0
    errors: List[str] = []

    for i, row in df.iterrows():
        try:
            stud_id = _safe_str(row.get("STUD_ID"))
            stud_name = _safe_str(row.get("STUD_FULL_NAME"))
            course = _safe_str(row.get("COURSE_CODE"))
            if not stud_id or not course:
                skipped += 1
                continue

            student = (
                db.query(models.Student)
                .filter(models.Student.student_code == stud_id)
                .first()
            )
            if not student:
                student = models.Student(student_code=stud_id, name=stud_name or stud_id)
                db.add(student)
                db.flush()

            instructor = _safe_str(row.get("INSTRUCTOR"))
            section = _safe_str(row.get("SECTION"))

            # dedupe by (student, course, instructor, section)
            existing = (
                db.query(models.FxRequest)
                .filter(
                    models.FxRequest.student_id == student.id,
                    models.FxRequest.course_code == course,
                    models.FxRequest.instructor == instructor,
                    models.FxRequest.section == section,
                )
                .first()
            )
            if existing:
                skipped += 1
                continue

            req = models.FxRequest(
                student_id=student.id,
                course_code=course,
                course_title=_safe_str(row.get("COURSE_TITLE")),
                instructor=instructor,
                section=section,
                faculty=_safe_str(row.get("FACULTY")),
                cipher=_safe_str(row.get("CIPHER")),
                speciality=_safe_str(row.get("SPECIALITY")),
                ects=_safe_int(row.get("ECTS")),
                course_year="",
            )
            db.add(req)
            added += 1
        except Exception as e:
            errors.append(f"Жол {i+2}: {e}")
    db.commit()
    return added, skipped, errors[:20]


def export_fx_excel(db: Session) -> bytes:
    """Generate 2-sheet BS-style FX schedule Excel."""
    wb = Workbook()

    # ---- Sheet 1: FX кестесі ----
    ws1 = wb.active
    ws1.title = "FX кестесі"

    ws1.cell(row=2, column=11, value='"БЕКІТЕМІН"\nSDU Бизнес мектебі деканы\n_______________Сабденалиев Б.\n"____"____________2025ж.').alignment = Alignment(wrap_text=True)
    ws1.cell(row=3, column=1, value="SDU БИЗНЕС МЕКТЕБІ")
    ws1.cell(row=4, column=1, value="6В04101 - ЭКОНОМИКА, 6В04102 - МЕНЕДЖМЕНТ, 6В04103 - ЕСЕП ЖӘНЕ АУДИТ,")
    ws1.cell(row=5, column=1, value="6В04104 - ҚАРЖЫ, 6В04105 - ДИДЖИТАЛ МАРКЕТИНГ БІЛІМ БЕРУ БАҒДАРЛАМАЛАРЫ")
    ws1.cell(row=6, column=1, value="АРАЛЫҚ АТТЕСТАТТАУ КЕСТЕСІ")
    ws1.cell(row=7, column=1, value="2024-2025 оқу жылы, көктемгі семестр FX")

    headers1 = [
        "Курс", "Мерзімі", "Уақыты", "Емтихан қабылдаушы",
        "Пән коды", "Пән атауы", "БББ атауы",
        "ECTS", "Студент саны", "Аудитория / Платфо*рма",
        "Емтихан форматы тест/проект, жазбаша/ ауызша", "Емтихан ұзақтығы",
    ]
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill = PatternFill("solid", fgColor="E5E7EB")
    for ci, h in enumerate(headers1, start=1):
        c = ws1.cell(row=9, column=ci, value=h)
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border
        c.fill = fill

    fx_exams = (
        db.query(models.FxExam)
        .order_by(models.FxExam.exam_date, models.FxExam.exam_time)
        .all()
    )
    for idx, ex in enumerate(fx_exams, start=10):
        time_str = f"{ex.exam_time.hour:02d}.{ex.exam_time.minute:02d}" if ex.exam_time else ""
        row_vals = [
            ex.course_year or "",
            ex.exam_date.strftime("%Y-%m-%d") if ex.exam_date else "",
            time_str,
            ex.instructor or "",
            ex.course_code or "",
            ex.course_name or "",
            ex.program_name or "",
            ex.ects if ex.ects is not None else "",
            len(ex.student_assignments),
            ex.room.room_number if ex.room else "",
            ex.exam_format or "",
            f"{ex.duration} мин" if ex.duration else "",
        ]
        for ci, v in enumerate(row_vals, start=1):
            c = ws1.cell(row=idx, column=ci, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border

    widths1 = [10, 12, 10, 24, 14, 24, 28, 6, 10, 18, 28, 12]
    for ci, w in enumerate(widths1, start=1):
        ws1.column_dimensions[ws1.cell(row=9, column=ci).column_letter].width = w

    # ---- Sheet 2: Өтініш берген білімгерлер тізім ----
    ws2 = wb.create_sheet(title="Өтініш берген білімгерлер тізім")

    headers2 = [
        "№", "FACULTY", "CIPHER", "SPECIALITY", "COURSE_CODE", "COURSE_TITLE",
        "SECTION", "INSTRUCTOR", "STUD_ID", "STUD_FULL_NAME",
        "КҮНІ", "САҒАТ", "Аудитория",
    ]
    for ci, h in enumerate(headers2, start=2):  # start col 2 to mirror layout
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # build a lookup: (student_id, course_code, instructor, section) -> FxExam
    sched_lookup = {}
    for ex in fx_exams:
        for sa in ex.student_assignments:
            sched_lookup[(sa.student_id, ex.course_code, ex.instructor or "", ex.section or "")] = ex

    requests = db.query(models.FxRequest).all()
    # sort: by faculty, course_code, student_name
    requests.sort(key=lambda r: (r.faculty or "", r.course_code, r.student.name if r.student else ""))

    for idx, req in enumerate(requests, start=2):
        ex = sched_lookup.get(
            (req.student_id, req.course_code, req.instructor or "", req.section or "")
        )
        date_str = ex.exam_date.strftime("%Y-%m-%d") if ex and ex.exam_date else ""
        time_str = f"{ex.exam_time.hour:02d}.{ex.exam_time.minute:02d}" if ex and ex.exam_time else ""
        room_str = ex.room.room_number if ex and ex.room else ""
        row_vals = [
            idx - 1,  # №
            req.faculty or "",
            req.cipher or "",
            req.speciality or "",
            req.course_code,
            req.course_title or "",
            req.section or "",
            req.instructor or "",
            req.student.student_code if req.student else "",
            req.student.name if req.student else "",
            date_str,
            time_str,
            room_str,
        ]
        for ci, v in enumerate(row_vals, start=2):
            c = ws2.cell(row=idx, column=ci, value=v)
            c.alignment = Alignment(vertical="top")
            c.border = border

    widths2 = [4, 4, 10, 12, 16, 14, 32, 8, 32, 14, 24, 18, 8, 12]
    for ci, w in enumerate(widths2, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
